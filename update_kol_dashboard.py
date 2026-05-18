#!/usr/bin/env python3
"""Loho House - KOL TikTok Dashboard Updater
Download Google Sheets KOL_TikTok_Database_PerMonth -> parse -> generate HTML.
"""
import sys, os, json, subprocess
from datetime import datetime
from openpyxl import load_workbook

SHEET_ID = "1tXktWbmjaShXx3yGaErQ_KazDhr0pp9d0FwtAckjSzg"
URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
OUT = sys.argv[1] if len(sys.argv) > 1 else "Dashboard_KOL_TikTok.html"
TMP = "/tmp/kol_dashboard.xlsx"


def parse_money(v):
    if v is None: return 0
    if isinstance(v,(int,float)): return float(v)
    s = str(v).replace('₫','').replace('đ','').replace(' ','').replace('.','').replace(',','.')
    try: return float(s)
    except: return 0


def parse_dt(s, sheet_month=None):
    """Parse date robust với cả D/M/Y và M/D/Y format."""
    if not s: return None
    if isinstance(s, datetime):
        dt = s
    else:
        s = str(s).strip()
        dt = None
        for fmt in ['%d/%m/%Y %H:%M:%S','%d/%m/%Y %H:%M','%d/%m/%Y','%m/%d/%Y %H:%M:%S','%m/%d/%Y %H:%M','%m/%d/%Y','%Y-%m-%d %H:%M:%S','%Y-%m-%d']:
            try:
                dt = datetime.strptime(s, fmt)
                break
            except: continue
        if dt is None: return None

    if sheet_month is not None and dt.month != sheet_month:
        try:
            swapped = dt.replace(month=dt.day, day=dt.month)
            if swapped.month == sheet_month:
                return swapped
        except ValueError:
            pass
        try:
            return dt.replace(month=sheet_month)
        except ValueError:
            return dt.replace(month=sheet_month, day=min(dt.day, 28))
    return dt


def main():
    print("Tai Google Sheets...")
    r = subprocess.run(["curl","-sL",URL,"-o",TMP], capture_output=True)
    if r.returncode != 0 or not os.path.exists(TMP) or os.path.getsize(TMP) < 1000:
        print("Loi tai file"); sys.exit(1)

    wb = load_workbook(TMP, read_only=True, data_only=True)
    months = sorted({int(s[1:].split('_')[0]) for s in wb.sheetnames if s.startswith('T') and '_DonHang' in s})

    orders = []
    for m in months:
        for r in wb[f'T{m}_DonHang'].iter_rows(min_row=2, values_only=True):
            if not r[0]: continue
            dt = parse_dt(r[28], sheet_month=m)
            orders.append({
                'id': str(r[0]),
                'product': r[2] or '',
                'qty': int(r[7] or 0),
                'payment': parse_money(r[5]),
                'status': r[10] or '',
                'kol': r[11] or '',
                'ct': r[12] or '',
                'com': parse_money(r[22]),
                'date': dt.strftime('%Y-%m-%d') if dt else '',
                'month': m,
            })

    print(f"{len(orders)} don tu {len(months)} thang")

    here = os.path.dirname(os.path.abspath(__file__))
    tpl_path = os.path.join(here, 'kol_dashboard_template.html')
    with open(tpl_path, 'r', encoding='utf-8') as f:
        tpl = f.read()

    out_path = OUT if os.path.isabs(OUT) else os.path.join(here, OUT)
    html = (tpl
            .replace('"__DATA__"', json.dumps(orders, ensure_ascii=False, separators=(',',':')))
            .replace('"__MONTHS__"', json.dumps(months))
            .replace('__LAST_UPDATE__', datetime.now().strftime('%Y-%m-%d %H:%M')))
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"OK: {out_path} ({os.path.getsize(out_path):,} bytes)")
    try: os.remove(TMP)
    except: pass


if __name__ == "__main__":
    main()
